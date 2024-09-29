from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image
import pytesseract
import tabula
from twilio.rest import Client  # Import Twilio Client

app = Flask(__name__)
CORS(app)

# Twilio Credentials
TWILIO_ACCOUNT_SID = 'AC700dcbab06f7d68acd463064417e61d1'  # Replace with your Account SID
TWILIO_AUTH_TOKEN = '3676a3eec24ef60971603008cb64ba01'      # Replace with your Auth Token
TWILIO_WHATSAPP_NUMBER = 'whatsapp:+14155238886'  # Replace with your Twilio WhatsApp number

# Initialize Twilio client
twilio_client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

UPLOAD_FOLDER = 'frontend/upload'
ALLOWED_EXTENSIONS = {'pdf', 'jpeg', 'jpg', 'png', 'xlsx'}

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_pdf(file_path):
    dfs = tabula.read_pdf(file_path, pages="all", multiple_tables=True)
    if not dfs:
        return None
    return pd.concat(dfs, ignore_index=True)

def process_image(file_path):
    try:
        img = Image.open(file_path)
        text = pytesseract.image_to_string(img)
        data = [line.split() for line in text.splitlines() if line.strip()]
        return pd.DataFrame(data)
    except Exception as e:
        print(f"Error processing image: {e}")
        return None

def process_excel(file_path):
    try:
        excel_data = pd.read_excel(file_path, sheet_name=None, header=7)
        return pd.concat(excel_data.values(), ignore_index=True)
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        return None

def read_input_file(file_path):
    file_extension = os.path.splitext(file_path)[1].lower()
    if file_extension == '.pdf':
        return process_pdf(file_path)
    elif file_extension in ['.jpeg', '.jpg', '.png']:
        return process_image(file_path)
    elif file_extension == '.xlsx':
        return process_excel(file_path)
    else:
        return None

def send_whatsapp_message(to, message):
    try:
        message = twilio_client.messages.create(
            body=message,
            from_=TWILIO_WHATSAPP_NUMBER,
            to=f'whatsapp:{to}'
        )
        return message.sid
    except Exception as e:
        print(f"Failed to send message: {e}")
        return None

@app.route('/process', methods=['POST'])
def process_file():
    if 'file' not in request.files:
        return jsonify({'message': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'No selected file'}), 400

    if not allowed_file(file.filename):
        return jsonify({'message': 'Invalid file type'}), 400
    
    filename = file.filename
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(file_path)

    combined_df = read_input_file(file_path)
    if combined_df is None:
        return jsonify({'message': 'No valid data found in the file'}), 400
    
    otp = "output_combined.xlsx"
    with pd.ExcelWriter(otp, engine='openpyxl') as writer:
        combined_df.to_excel(writer, sheet_name="Combined_Data", index=False, startrow=4)
    
    wb = load_workbook(otp)
    ws = wb['Combined_Data']
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    student_name_column = 1
    parent_contact_column = 2  # Assuming parents' contacts are in the second column
    header_row_subjects = 7

    highlighted = False
    parents_notified = []
    for row in ws.iter_rows(min_row=header_row_subjects + 1):
        attendance_cell = row[0]  # Assuming attendance is in the first column
        try:
            if isinstance(attendance_cell.value, str):
                attendance_value = float(attendance_cell.value.replace('%', '').replace(',', '').strip())
            else:
                attendance_value = float(attendance_cell.value)

            if attendance_value < 60:
                student_name_cell = ws.cell(row=attendance_cell.row, column=student_name_column)
                student_name_cell.fill = highlight_fill
                attendance_cell.fill = highlight_fill
                highlighted = True
                
                # Get parent's phone number and send a message
                parent_contact = row[parent_contact_column].value
                if parent_contact not in parents_notified:
                    message = f"Dear Parent, your child {student_name_cell.value} has less than 60% attendance."
                    send_whatsapp_message(parent_contact, message)
                    parents_notified.append(parent_contact)
        except ValueError:
            pass
    
    output_file = "output_highlighted.xlsx"
    wb.save(output_file)

    return jsonify({
        'message': f'Highlighting completed. Check {output_file}',
        'file_link': f'/download/{output_file}'
    }), 200

@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    return send_file(filename, as_attachment=True)

# Directly send a WhatsApp message for short attendance to the number 9589497927
@app.route('/send-whatsapp', methods=['GET'])
def send_whatsapp():
    to_number = '919589497927'  # Number to send the message to, in international format
    message = "Short Attendance"
    
    # Sending the message
    msg_id = send_whatsapp_message(to_number, message)
    
    if msg_id:
        return jsonify({'message': 'WhatsApp message sent successfully!', 'msg_id': msg_id}), 200
    else:
        return jsonify({'message': 'Failed to send WhatsApp message.'}), 500

if __name__ == '__main__':
    app.run(port=5000, debug=True)